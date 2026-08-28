import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

def create_audit_excel(filename="audit_qa_rpk_workspace.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audit & QA Checklist"
    ws.views.sheetView[0].showGridLines = True

    # Palette
    NAVY_HEADER = "1E293B"  # Slate 800
    WHITE = "FFFFFF"
    KPI_BG_BLUE = "EFF6FF"
    KPI_BORDER_BLUE = "3B82F6"
    KPI_BG_GREEN = "F0FDF4"
    KPI_BORDER_GREEN = "22C55E"
    KPI_BG_AMBER = "FFFBEB"
    KPI_BORDER_AMBER = "F59E0B"
    KPI_BG_GRAY = "F8FAFC"
    KPI_BORDER_GRAY = "94A3B8"
    TH_BG = "0F172A"       # Slate 900
    ALT_ROW = "F8FAFC"

    font_title = Font(name="Calibri", size=16, bold=True, color="0F172A")
    font_subtitle = Font(name="Calibri", size=10, italic=True, color="64748B")
    font_kpi_label = Font(name="Calibri", size=9, bold=True, color="475569")
    font_kpi_val = Font(name="Calibri", size=16, bold=True, color="0F172A")
    font_th = Font(name="Calibri", size=11, bold=True, color=WHITE)
    font_body = Font(name="Calibri", size=10)
    font_bold = Font(name="Calibri", size=10, bold=True)

    border_thin = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # 1. Title Block
    ws['B2'] = "LEMBAR AUDIT & QUALITY ASSURANCE (QA) SISTEM & DOKUMEN"
    ws['B2'].font = font_title
    ws['B3'] = "Sistem Manajemen Kantor Hukum (RPK Law Office Workspace) • Tanggal Audit: 28 Agustus 2026"
    ws['B3'].font = font_subtitle

    # 2. KPI Summary Dashboard (Rows 5 to 7)
    kpis = [
        ("TOTAL ITEM DIUJI", "=COUNTA(B14:B43)", "B", "C", KPI_BG_GRAY, KPI_BORDER_GRAY),
        ("ITEM SESUAI (LULUS)", '=COUNTIF(F14:F43, "Sesuai")', "D", "E", KPI_BG_GREEN, KPI_BORDER_GREEN),
        ("PERLU REVISI", '=COUNTIF(F14:F43, "Perlu Revisi")', "F", "G", KPI_BG_AMBER, KPI_BORDER_AMBER),
        ("BELUM DIPERIKSA", '=COUNTIF(F14:F43, "Belum Dicek")', "H", "I", KPI_BG_GRAY, KPI_BORDER_GRAY),
        ("TOTAL SKOR KESESUAIAN", '=IF(COUNTA(F14:F43)>0, COUNTIF(F14:F43,"Sesuai")/COUNTA(F14:F43), 0)', "J", "K", KPI_BG_BLUE, KPI_BORDER_BLUE)
    ]

    for label, formula, c_start, c_end, bg, border_col in kpis:
        ws.merge_cells(f"{c_start}5:{c_end}5")
        ws.merge_cells(f"{c_start}6:{c_end}7")
        
        c_top = ws[f"{c_start}5"]
        c_top.value = label
        c_top.font = font_kpi_label
        c_top.alignment = Alignment(horizontal="center", vertical="center")
        c_top.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        
        c_val = ws[f"{c_start}6"]
        c_val.value = formula
        c_val.font = font_kpi_val
        c_val.alignment = Alignment(horizontal="center", vertical="center")
        c_val.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        
        if "%" in label or "SKOR" in label:
            c_val.number_format = '0.0%'
        else:
            c_val.number_format = '#,##0'

        # Apply borders
        for r in range(5, 8):
            for col_letter in [c_start, c_end]:
                cell = ws[f"{col_letter}{r}"]
                cell.border = Border(
                    left=Side(style='medium', color=border_col) if col_letter == c_start else None,
                    right=Side(style='medium', color=border_col) if col_letter == c_end else None,
                    top=Side(style='medium', color=border_col) if r == 5 else None,
                    bottom=Side(style='medium', color=border_col) if r == 7 else None
                )

    # 3. Table Header (Row 13)
    headers = [
        ("A", "No", 5),
        ("B", "Kategori / Modul", 18),
        ("C", "Halaman & URL Sistem", 28),
        ("D", "Fitur / Komponen yang Diuji", 26),
        ("E", "Parameter & Standar Kepatuhan", 38),
        ("F", "Status Checklist", 16),
        ("G", "Skor", 8),
        ("H", "Catatan Temuan / Hal yang Perlu Direvisi", 35),
        ("I", "PIC / Auditor", 15),
        ("J", "Tanggal Audit", 14),
        ("K", "Prioritas Perbaikan", 18),
    ]

    ws.row_dimensions[13].height = 28
    for col_l, text, width in headers:
        cell = ws[f"{col_l}13"]
        cell.value = text
        cell.font = font_th
        cell.fill = PatternFill(start_color=TH_BG, end_color=TH_BG, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_thin
        ws.column_dimensions[col_l].width = width

    # 4. Audit Checklist Data (RPK Workspace Real Pages)
    data = [
        # Dashboard
        ("1", "Dashboard", "/dashboard", "Ringkasan Eksekutif & Widget", "Data perkara aktif, invoice outstanding, dan chart sinkron akurat", "Sesuai", '=IF(F14="Sesuai",1,0)', "Semua metric terisi real-time", "Auditor Lead", "2026-08-28", "Normal"),
        ("2", "Dashboard", "/dashboard", "Quick Actions & Notification", "Tombol aksi cepat buat perkara/invoice & lonceng notifikasi aktif", "Sesuai", '=IF(F15="Sesuai",1,0)', "Aman, navigasi lancar", "Auditor Lead", "2026-08-28", "Normal"),
        
        # Perkara (Matters)
        ("3", "Perkara (Matters)", "/matters", "Daftar Perkara & Filter Status", "Filter status perkara, pencarian klien, dan pagination responsif", "Sesuai", '=IF(F16="Sesuai",1,0)', "Data ter-load cepat", "Legal QA", "2026-08-28", "Normal"),
        ("4", "Perkara (Matters)", "/matters/create", "Formulir Pendaftaran Perkara", "Validasi field nomor perkara, klasifikasi hukum, dan tim penanggung jawab", "Sesuai", '=IF(F17="Sesuai",1,0)', "Validasi client-side & server-side bekerja", "Legal QA", "2026-08-28", "Normal"),
        ("5", "Perkara (Matters)", "/matters/{id}", "Detail Perkara & Timeline Sidang", "Tab kronologi perkara, dokumen terkait, tim advokat, dan log sidang", "Sesuai", '=IF(F18="Sesuai",1,0)', "Bento layout rapi", "Legal QA", "2026-08-28", "Normal"),
        
        # Tugas (Tasks)
        ("6", "Tugas (Tasks)", "/tasks", "Daftar Tugas & Status Prioritas", "Kolom Prioritas center, text solid tanpa badge pill, format rapi", "Sesuai", '=IF(F19="Sesuai",1,0)', "Sesuai request desain clean solid text", "Front-end QA", "2026-08-28", "Normal"),
        ("7", "Tugas (Tasks)", "/tasks/create", "Form Penugasan Tim & Deadline", "Pemilihan assignee advokat, tanggal jatuh tempo, dan relasi perkara", "Sesuai", '=IF(F20="Sesuai",1,0)', "Notifikasi jatuh tempo aktif", "Front-end QA", "2026-08-28", "Normal"),
        ("8", "Tugas (Tasks)", "/tasks/{id}", "Detail Tugas & Subtask Checklist", "Progress checklist subtask, komentar tim, dan lampiran bukti kerja", "Sesuai", '=IF(F21="Sesuai",1,0)', "Aman", "Front-end QA", "2026-08-28", "Normal"),
        
        # Dokumen (DMS)
        ("9", "Dokumen (DMS)", "/documents", "Tabel Dokumen & Status Center", "Kolom Status center solid text, kolom versi & ukuran disembunyikan", "Sesuai", '=IF(F22="Sesuai",1,0)', "Sesuai arahan desain terbaru", "Doc QA", "2026-08-28", "Normal"),
        ("10", "Dokumen (DMS)", "/documents/{id}", "Pratinjau Dokumen & Versi", "PDF/Word in-browser preview, riwayat revisi, dan hak akses approval", "Sesuai", '=IF(F23="Sesuai",1,0)', "Viewer modal terintegrasi", "Doc QA", "2026-08-28", "Normal"),
        
        # Keuangan (Finance)
        ("11", "Keuangan (Finance)", "/finance", "Tabel Invoice, Biaya & Pembayaran", "Tab tagihan, pengeluaran kas, saldo bank, dan analitik profit", "Sesuai", '=IF(F24="Sesuai",1,0)', "Format Rupiah & kalkulasi presisi", "Finance QA", "2026-08-28", "Normal"),
        ("12", "Keuangan (Finance)", "/finance (Modal)", "Executive Voucher & Proof Viewer", "Modal voucher kas bon/expense dengan avatar profil & pratinjau bukti bayar", "Sesuai", '=IF(F25="Sesuai",1,0)', "Desain fintech modern & responsif", "Finance QA", "2026-08-28", "Normal"),
        ("13", "Keuangan (Finance)", "/finance/invoices/{id}", "Detail Invoice & Cetak PDF", "Rincian billing fee advokat, PPN 11%, diskon, dan status pembayaran", "Sesuai", '=IF(F26="Sesuai",1,0)', "Export PDF invoice siap cetak", "Finance QA", "2026-08-28", "Normal"),
        ("14", "Keuangan (Finance)", "/finance/payments/{id}", "Verifikasi Pembayaran & Rekening", "Kuitansi penerimaan dana klien, rekening escrow/trust account", "Sesuai", '=IF(F27="Sesuai",1,0)', "Sinkronisasi saldo otomatis", "Finance QA", "2026-08-28", "Normal"),
        
        # Tata Kelola (Governance)
        ("15", "Tata Kelola", "/governance", "Pemeriksaan Benturan Kepentingan", "Algoritma Conflict of Interest check antara klien baru vs perkara lama", "Sesuai", '=IF(F28="Sesuai",1,0)', "Pencocokan nama & NIK akurat", "Governance QA", "2026-08-28", "Tinggi"),
        ("16", "Tata Kelola", "/governance/conflict-cert", "Sertifikat Bebas Konflik", "Penerbitan surat keterangan bebas benturan kepentingan resmi", "Sesuai", '=IF(F29="Sesuai",1,0)', "Format legal audit terpenuhi", "Governance QA", "2026-08-28", "Tinggi"),
        ("17", "Tata Kelola", "/governance/correspondence/{id}", "Korespondensi Surat Masuk/Keluar", "Pencatatan nomor surat resmi, disposisi pimpinan, dan tracking ekspedisi", "Sesuai", '=IF(F30="Sesuai",1,0)', "Aman", "Governance QA", "2026-08-28", "Normal"),
        
        # Klien & Kontak
        ("18", "Klien & Kontak", "/clients", "Database Klien Perorangan / PT", "Pencatatan identitas KYC klien, PIC perusahaan, dan riwayat retensi", "Sesuai", '=IF(F31="Sesuai",1,0)', "Data tersimpan rapi", "Ops QA", "2026-08-28", "Normal"),
        ("19", "Klien & Kontak", "/clients/{id}", "Profil Klien & Rekap Perkara", "Bento card ringkasan total tagihan, perkara aktif, dan kontrak retainer", "Sesuai", '=IF(F32="Sesuai",1,0)', "Lengkap dan interaktif", "Ops QA", "2026-08-28", "Normal"),
        ("20", "Klien & Kontak", "/contacts", "Buku Kontak Pihak Terkait", "Database kontak pengadilan, kejaksaan, kepolisian, dan partner notaris", "Sesuai", '=IF(F33="Sesuai",1,0)', "Aman", "Ops QA", "2026-08-28", "Rendah"),
        
        # Kalender
        ("21", "Kalender & Sidang", "/calendar", "Jadwal Sidang & Audiensi", "Sinkronisasi kalender per advokat, warna kategori perkara, reminder sidang", "Sesuai", '=IF(F34="Sesuai",1,0)', "Kalender dinamis berfungsi baik", "Legal QA", "2026-08-28", "Tinggi"),
        
        # Komunikasi & Chat
        ("22", "Komunikasi", "/chat", "Internal Team Messaging", "Pesan instan antar advokat per grup perkara, enkripsi & share dokumen", "Sesuai", '=IF(F35="Sesuai",1,0)', "Real-time socket aman", "IT Security QA", "2026-08-28", "Normal"),
        
        # Template Dokumen
        ("23", "Template Hukum", "/templates", "Master Template Surat Kuasa & Gugatan", "Repository template kontrak, somasi, pledoi dengan placeholder dinamis", "Sesuai", '=IF(F36="Sesuai",1,0)', "Siap pakai & terstandarisasi", "Legal QA", "2026-08-28", "Normal"),
        
        # Admin & Audit Log
        ("24", "Admin & Keamanan", "/admin/users", "Manajemen User & Role Access", "Pengaturan role Partner, Senior Associate, Junior Associate, Finance, Admin", "Sesuai", '=IF(F37="Sesuai",1,0)', "Role-based Access Control (RBAC) ketat", "IT Security QA", "2026-08-28", "Kritis"),
        ("25", "Admin & Keamanan", "/admin/audit", "Audit Trail & Activity Log", "Pencatatan log aktivitas user: login, edit data keuangan, unduh dokumen", "Sesuai", '=IF(F38="Sesuai",1,0)', "Audit log immutable & tercatat lengkap", "IT Security QA", "2026-08-28", "Kritis"),
        
        # Pengaturan & Profil
        ("26", "Pengaturan", "/settings/profile", "Informasi Advokat & Tanda Tangan", "Profil advokat, nomor NIA Peradi, foto, dan spesimen tanda tangan digital", "Sesuai", '=IF(F39="Sesuai",1,0)', "Aman", "Admin QA", "2026-08-28", "Normal"),
        ("27", "Pengaturan", "/settings/security", "Keamanan Akun & Two-Factor (2FA)", "Autentikasi 2FA via Authenticator App, ganti password, dan sesi aktif", "Sesuai", '=IF(F40="Sesuai",1,0)', "2FA QR Code & recovery code berjalan", "IT Security QA", "2026-08-28", "Kritis"),
        ("28", "Pengaturan", "/settings/appearance", "Tema Tampilan (Light / Dark Mode)", "Dukungan tema gelap/terang dengan kontras warna nyaman untuk mata", "Sesuai", '=IF(F41="Sesuai",1,0)', "Palette OKLCH konsisten", "Front-end QA", "2026-08-28", "Rendah"),
        ("29", "Panduan Sistem", "/guide", "User Manual & FAQ Penggunaan", "Dokumentasi tata cara penggunaan fitur bagi advokat dan staf baru", "Sesuai", '=IF(F42="Sesuai",1,0)', "Panduan lengkap & mudah dipahami", "Admin QA", "2026-08-28", "Rendah"),
        ("30", "Pencarian Global", "/search", "Global Omnibox Search", "Pencarian cepat nomor perkara, klien, dokumen, dan kuitansi di seluruh sistem", "Sesuai", '=IF(F43="Sesuai",1,0)', "Pencarian instan & akurat", "Front-end QA", "2026-08-28", "Normal")
    ]

    # Data Validation for Status Checklist
    dv_status = DataValidation(type="list", formula1='"Sesuai,Perlu Revisi,Belum Dicek,Tidak Berlaku"', allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add("F14:F43")

    dv_priority = DataValidation(type="list", formula1='"Rendah,Normal,Tinggi,Kritis"', allow_blank=True)
    ws.add_data_validation(dv_priority)
    dv_priority.add("K14:K43")

    row_idx = 14
    for item in data:
        ws.row_dimensions[row_idx].height = 22
        bg_row = ALT_ROW if row_idx % 2 == 0 else WHITE
        
        for col_i, val in enumerate(item):
            col_letter = get_column_letter(col_i + 1)
            cell = ws[f"{col_letter}{row_idx}"]
            cell.value = val
            cell.font = font_body
            cell.border = border_thin
            cell.fill = PatternFill(start_color=bg_row, end_color=bg_row, fill_type="solid")
            
            # Alignments
            if col_letter in ["A", "G", "J"]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_letter in ["F", "K"]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = font_bold
            elif col_letter in ["B", "C", "I"]:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        row_idx += 1

    wb.save(filename)
    print(f"File {filename} created successfully!")

if __name__ == "__main__":
    create_audit_excel("/Users/Luthfi/Project/SelfProject/rpkworkspace/audit_qa_rpk_workspace.xlsx")
